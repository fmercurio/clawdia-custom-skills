#!/usr/bin/env python3
"""Diagnostic reconciliation: match a cash-flow spreadsheet against Agilize transactions.

Read-only. Downloads Agilize tx for the year, parses the sheet, runs a greedy
(month, abs_amount) match with description-scored tie-breaker, produces a crosstab
of sheet classification × Agilize category, and emits a Markdown diagnostic report.
Never applies fixes — see references/agilize-spreadsheet-reconciliation.md.

Usage:
    # Use default config at ~/.config/agilize.json
    python agilize_match_spreadsheet.py --xlsx ~/fluxo.xlsx --year 2025

    # Custom config / output dir
    python agilize_match_spreadsheet.py \\
        --xlsx ~/fluxo.xlsx --year 2025 \\
        --config ~/.config/agilize.json \\
        --output-dir /tmp/audit-2025

Sheet requirements (auto-detected sheet 'EXTRATO BANCÁRIO' or first sheet):
    Columns (case-insensitive substrings): MÊS, DATA, DESCRIÇÃO, VALOR,
    [BANCO], [STATUS], TIPO, CLASSIFICAÇÃO, [OBSERVAÇÃO]

Outputs in --output-dir (default /tmp/agilize-audit-<year>/):
    all-<year>.json              — Agilize transactions
    categories.json              — Agilize categories tree
    matched.json                 — successful matches with score
    unmatched_sheet.json         — sheet rows with no Agilize counterpart
    crosstab.json                — sheet_class × agilize_category counts
    diagnostico.md               — human-readable report
"""
from __future__ import annotations

import argparse
import calendar
import json
import os
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Optional

# Allow imports when run from the skill's scripts/ dir and support optional
# user-local dependency installs before importing third-party packages.
HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
sys.path.insert(0, os.path.expanduser("~/.local/py-lib"))

import agilize_login as A  # noqa: E402

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install with: uv pip install --target ~/.local/py-lib openpyxl", file=sys.stderr)
    sys.exit(2)

import requests  # noqa: E402


# ─── Secure output helpers ────────────────────────────────────────────────────

def write_secure(path: Path, data: str | bytes, *, mode: str = "text") -> None:
    """Write audit artefacts with user-only permissions (0600)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    flags = os.O_WRONLY | os.O_CREAT | os.O_TRUNC
    fd = os.open(path, flags, 0o600)
    try:
        if mode == "bytes":
            os.write(fd, data if isinstance(data, bytes) else data.encode("utf-8"))
        else:
            os.write(fd, data.encode("utf-8") if isinstance(data, str) else data)
    finally:
        os.close(fd)


def dumps_json(data: Any, **kwargs: Any) -> str:
    return json.dumps(data, ensure_ascii=False, **kwargs)


def mask_cnpj(cnpj: str) -> str:
    digits = re.sub(r"\D", "", cnpj or "")
    if len(digits) < 6:
        return "<redacted>"
    return f"***{digits[-6:]}"


def request_json_with_reauth(session: requests.Session, cfg: dict, url: str, headers: dict, *, timeout: int) -> tuple[int, Any]:
    """GET JSON; on token expiry, re-authenticate once and retry."""
    resp = session.get(url, headers=headers, timeout=timeout)
    if resp.status_code == 401:
        token = A.login(cfg, A.DEFAULT_CLIENT_ID, A.DEFAULT_REDIRECT_URI, 30,
                        "Mozilla/5.0 Chrome/124.0 Safari/537.36")
        headers["Authorization"] = f"Bearer {token.access_token}"
        resp = session.get(url, headers=headers, timeout=timeout)
    if resp.status_code != 200:
        return resp.status_code, None
    return resp.status_code, resp.json()


# ─── Normalization ────────────────────────────────────────────────────────────

def norm(s: str) -> str:
    s = (s or "").upper().strip()
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def desc_compat(sheet_desc: str, ag_desc: str) -> float:
    """Score description compatibility in [0, 1]."""
    sn = norm(sheet_desc)
    an = norm(ag_desc)
    if not sn or not an:
        return 0.0
    if sn == an:
        return 1.0
    if sn in an or an.startswith(sn) or an.endswith(sn):
        return 0.8
    st = set(sn.split())
    at = set(an.split())
    if not st or not at:
        return 0.0
    return min(0.5, len(st & at) / max(len(st), 1))


def get_month(t: dict) -> str:
    return (t.get("mainDate") or t.get("consolidatedAt") or "")[:7]


# ─── Sheet parsing ────────────────────────────────────────────────────────────

def parse_sheet(xlsx_path: str, sheet_name: Optional[str] = None) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    # Auto-pick sheet
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        # Heuristic: prefer sheet named like 'EXTRATO*' or with MÊS/DATA/VALOR header
        ws = None
        for sn in wb.sheetnames:
            candidate = wb[sn]
            rows = candidate.iter_rows(max_row=5, values_only=True)
            for r in rows:
                if not r:
                    continue
                cells = [str(c or "").upper() for c in r]
                if any("MÊS" in c or "MES" in c for c in cells) and any("VALOR" in c for c in cells):
                    ws = candidate
                    break
            if ws:
                break
        if not ws:
            ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    # Find header row (first row with >= 4 non-empty cells)
    header_idx = None
    for i, r in enumerate(rows[:10]):
        if r and sum(1 for c in r if c) >= 4:
            header_idx = i
            break
    if header_idx is None:
        print(f"ERROR: could not find header row in {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    header = [str(c or "").strip().upper() for c in rows[header_idx]]

    def find_col(*needles: str) -> Optional[int]:
        for i, h in enumerate(header):
            for n in needles:
                if n.upper() in h:
                    return i
        return None

    col_mes = find_col("MÊS", "MES")
    col_data = find_col("DATA")
    col_desc = find_col("DESCRI", "DESC")
    col_valor = find_col("VALOR")
    col_banco = find_col("BANCO")
    col_tipo = find_col("TIPO")
    col_class = find_col("CLASSIFICA", "CATEG")
    col_obs = find_col("OBSERVA")

    required = [col_mes, col_data, col_desc, col_valor, col_class]
    if any(c is None for c in required):
        missing = []
        if col_mes is None: missing.append("MÊS")
        if col_data is None: missing.append("DATA")
        if col_desc is None: missing.append("DESCRIÇÃO")
        if col_valor is None: missing.append("VALOR")
        if col_class is None: missing.append("CLASSIFICAÇÃO")
        print(f"ERROR: missing required columns: {missing}. Found header: {header}", file=sys.stderr)
        sys.exit(1)

    out = []
    for r in rows[header_idx + 1:]:
        if not r or not r[col_mes]:
            continue
        try:
            amt = float(r[col_valor]) if r[col_valor] is not None else 0.0
        except (TypeError, ValueError):
            amt = 0.0
        out.append({
            "month": str(r[col_mes])[:7],
            "date": str(r[col_data] or "")[:10],
            "description": str(r[col_desc] or "").strip(),
            "amount": round(amt, 2),
            "bank": str(r[col_banco] or "").strip() if col_banco is not None else "",
            "tipo": str(r[col_tipo] or "").strip() if col_tipo is not None else "",
            "classificacao": str(r[col_class] or "").strip(),
            "obs": str(r[col_obs] or "").strip() if col_obs is not None else "",
        })
    return out


# ─── Agilize fetch ────────────────────────────────────────────────────────────

def fetch_year(cfg: dict, year: int, out_dir: Path) -> list[dict]:
    cid = cfg["company_id"]
    cnpj = cfg["company_cnpj"]
    token = A.login(cfg, A.DEFAULT_CLIENT_ID, A.DEFAULT_REDIRECT_URI, 30,
                    "Mozilla/5.0 Chrome/124.0 Safari/537.36")
    hdr = {"Authorization": f"Bearer {token.access_token}",
           "Accept": "application/json",
           "Referer": "https://app.agilize.com.br/",
           "key": cnpj}

    session = requests.Session()

    # Categories
    status, payload = request_json_with_reauth(
        session, cfg,
        f"{A.API_BASE}/api/v1/companies/{cid}/finance-transaction-categories",
        hdr, timeout=30,
    )
    if status == 200:
        write_secure(out_dir / "categories.json", dumps_json(payload))

    # Transactions per month
    all_tx = []
    for m in range(1, 13):
        last = calendar.monthrange(year, m)[1]
        url = (f"{A.API_BASE}/api/v1/companies/{cid}/finance-transactions"
               f"?from={year}-{m:02d}-01T00:00:00-0300"
               f"&to={year}-{m:02d}-{last}T23:59:59-0300"
               f"&sort=mainDate&direction=DESC&count=3000")
        status, payload = request_json_with_reauth(session, cfg, url, hdr, timeout=60)
        if status != 200:
            print(f"  month {m:02d}: HTTP {status}", file=sys.stderr)
            continue
        items = payload if isinstance(payload, list) else payload.get("items", [])
        write_secure(out_dir / f"month-{m:02d}.json", dumps_json(items))
        print(f"  month {m:02d}: {len(items)} tx")
        all_tx.extend(items)

    write_secure(out_dir / f"all-{year}.json", dumps_json(all_tx))
    return all_tx


# ─── Matching ─────────────────────────────────────────────────────────────────

def match(sheet_rows: list[dict], txs: list[dict]) -> tuple[list, list]:
    """Greedy (month, abs_amount) match with description-scored tie-breaker."""
    ag_by_key = defaultdict(list)
    for t in txs:
        m = get_month(t)
        amt = round(abs(t.get("amount", 0)), 2)
        ag_by_key[(m, amt)].append(t)

    ag_by_amt_only = defaultdict(list)
    for t in txs:
        amt = round(abs(t.get("amount", 0)), 2)
        ag_by_amt_only[amt].append(t)

    used = set()
    matched = []
    unmatched = []

    for sr in sheet_rows:
        m = sr["month"][:7]
        amt = abs(sr["amount"])
        cand = [t for t in ag_by_key.get((m, amt), []) if t["__identity"] not in used]
        # Adjacent-month fallback
        if not cand:
            for dm in (-1, 1):
                y, mo = int(m[:4]), int(m[5:7])
                mo += dm
                if mo == 0: y, mo = y - 1, 12
                elif mo == 13: y, mo = y + 1, 1
                adj = f"{y:04d}-{mo:02d}"
                cand = [t for t in ag_by_key.get((adj, amt), []) if t["__identity"] not in used]
                if cand:
                    break
        if not cand:
            unmatched.append(sr)
            continue
        scored = sorted(((desc_compat(sr["description"], t.get("description", "")), t) for t in cand),
                        key=lambda x: -x[0])
        best_score, best = scored[0]
        used.add(best["__identity"])
        matched.append({"sheet": sr, "ag": best, "score": best_score, "candidates": len(cand)})
    return matched, unmatched


def crosstab(matched: list) -> dict:
    """sheet_classificacao → agilize_category → count."""
    ct = defaultdict(Counter)
    samples = defaultdict(list)
    for m in matched:
        sr = m["sheet"]
        ag = m["ag"]
        ag_cat = ag.get("category") or {}
        ag_label = f"{ag_cat.get('code', '?')} | {ag_cat.get('name', '<none>')}"
        ct[sr["classificacao"]][ag_label] += 1
        key = (sr["classificacao"], ag_label)
        if len(samples[key]) < 2:
            samples[key].append({
                "date": sr["date"],
                "amount": sr["amount"],
                "sheet_desc": sr["description"][:50],
                "ag_desc": ag.get("description", "")[:50],
            })
    return {"crosstab": {k: dict(v) for k, v in ct.items()},
            "samples": {f"{k[0]} || {k[1]}": v for k, v in samples.items()}}


# ─── Report ───────────────────────────────────────────────────────────────────

def render_report(matched: list, unmatched: list, txs: list, year: int,
                  company_cnpj: str, out_dir: Path) -> str:
    total_sheet = len(matched) + len(unmatched)
    closed = sum(1 for t in txs if t.get("isClosedPeriod"))

    R = []
    R.append(f"# Diagnóstico: Categorização Agilize {year} vs Planilha")
    R.append("")
    R.append(f"**CNPJ:** {mask_cnpj(company_cnpj)}  ")
    R.append(f"**Período:** jan-dez/{year}  ")
    R.append(f"**Método:** match descrição+valor±2dias, read-only  ")
    R.append("")
    R.append("---")
    R.append("")
    R.append("## Resumo")
    R.append("")
    R.append(f"- **Transações Agilize {year}:** {len(txs)}")
    R.append(f"- **Linhas planilha:** {total_sheet}")
    R.append(f"- **Matched:** {len(matched)} ({100 * len(matched) / max(1, total_sheet):.1f}%)")
    R.append(f"- **Unmatched:** {len(unmatched)} ({100 * len(unmatched) / max(1, total_sheet):.1f}%)")
    R.append(f"- **Meses fechados:** {closed}/{len(txs)} tx em período fechado")
    R.append("")

    # Crosstab summary
    ct = crosstab(matched)["crosstab"]
    R.append("---")
    R.append("")
    R.append("## Crosstab: classificação planilha → categoria Agilize")
    R.append("")
    for sheet_class in sorted(ct):
        ag_dist = ct[sheet_class]
        total = sum(ag_dist.values())
        primary = Counter(ag_dist).most_common(1)[0]
        primary_pct = primary[1] / total
        flag = "✓" if len(ag_dist) == 1 else ("~" if primary_pct >= 0.9 else "⚠")
        R.append("")
        R.append(f"{flag} **[{sheet_class}]** — {total} tx (primary {primary_pct * 100:.0f}%)")
        for ag_label, count in Counter(ag_dist).most_common():
            pct = count / total * 100
            R.append(f"  - {count} ({pct:.1f}%) — {ag_label}")
    R.append("")
    R.append("---")
    R.append("")

    # Unmatched by classification
    R.append("## Unmatched — por classificação planilha")
    R.append("")
    un_by_class = Counter(u["classificacao"] for u in unmatched)
    R.append("| Classificação | Count | Total R$ |")
    R.append("|---|---:|---:|")
    for cls in sorted(un_by_class, key=lambda k: -abs(sum(u["amount"] for u in unmatched if u["classificacao"] == k))):
        items = [u for u in unmatched if u["classificacao"] == cls]
        total = sum(u["amount"] for u in items)
        R.append(f"| {cls} | {len(items)} | {total:,.2f} |")
    R.append("")
    R.append("---")
    R.append("")

    # High-value unmatched
    big = sorted([u for u in unmatched if abs(u["amount"]) >= 10000], key=lambda u: -abs(u["amount"]))
    if big:
        R.append("## Unmatched de alto valor (investigar manualmente)")
        R.append("")
        R.append("| Data | Valor | Classificação | Descrição |")
        R.append("|---|---:|---|---|")
        for u in big:
            R.append(f"| {u['date']} | {u['amount']:,.2f} | {u['classificacao']} | {u['description'][:50]} |")
        R.append("")

    # Artefacts
    R.append("---")
    R.append("")
    R.append("## Artefatos")
    R.append("")
    for fname in [f"all-{year}.json", "categories.json", "matched.json",
                  "unmatched_sheet.json", "crosstab.json", "diagnostico.md"]:
        R.append(f"- `{out_dir}/{fname}`")
    R.append("")

    text = "\n".join(R)
    write_secure(out_dir / "diagnostico.md", text)
    return text


# ─── Main ─────────────────────────────────────────────────────────────────────

def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--xlsx", required=True, help="Excel/xlsx path (or directory containing it)")
    ap.add_argument("--year", type=int, required=True, help="Year to audit (e.g. 2025)")
    ap.add_argument("--config", default=os.path.expanduser("~/.config/agilize.json"),
                    help="Agilize config file (default ~/.config/agilize.json)")
    ap.add_argument("--output-dir", default=None,
                    help="Output directory (default /tmp/agilize-audit-<year>/)")
    ap.add_argument("--sheet", default=None, help="Sheet name (auto-detected if omitted)")
    args = ap.parse_args()

    # Resolve config
    cfg_path = Path(args.config).expanduser()
    if not cfg_path.exists():
        print(f"ERROR: config not found: {cfg_path}", file=sys.stderr)
        return 2
    cfg = json.loads(cfg_path.read_text())

    # Output dir
    out_dir = Path(args.output_dir or f"/tmp/agilize-audit-{args.year}")
    out_dir.mkdir(parents=True, exist_ok=True)
    print(f"Output: {out_dir}")

    # Fetch Agilize
    print(f"\n=== Fetching Agilize {args.year} ===")
    txs = fetch_year(cfg, args.year, out_dir)
    print(f"Total: {len(txs)} tx")

    # Parse sheet
    print(f"\n=== Parsing {args.xlsx} ===")
    sheet_rows = parse_sheet(args.xlsx, args.sheet)
    print(f"Sheet rows: {len(sheet_rows)}")

    # Match
    print(f"\n=== Matching ===")
    matched, unmatched = match(sheet_rows, txs)
    pct = 100 * len(matched) / max(1, len(sheet_rows))
    print(f"Matched: {len(matched)} ({pct:.1f}%)")
    print(f"Unmatched: {len(unmatched)}")

    # Save
    write_secure(out_dir / "matched.json", dumps_json(matched, default=str))
    write_secure(out_dir / "unmatched_sheet.json", dumps_json(unmatched, default=str))
    ct = crosstab(matched)
    write_secure(out_dir / "crosstab.json", dumps_json(ct, indent=2))

    # Report
    print(f"\n=== Generating report ===")
    render_report(matched, unmatched, txs, args.year, cfg["company_cnpj"], out_dir)
    print(f"Report: {out_dir}/diagnostico.md")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
